#!/usr/bin/python
# -*- coding: latin-1 -*-
"""
"""
import os
import sys
import time
import yaml
import pdb

from collections import OrderedDict
from openpyxl import load_workbook

ROW_KEYS = [
    'Number',
    'Title',
    'Surname',
    'Name',
    'Publisher',
    'City',
    'Year',
    'Category',
    'ISBN',
    'Borrowing',
    'Available', # not important at the moment
    'Location',
    'White card',
    'Book card',
    'Kind',
    'Amount',
    'Marking',
    'Anotations',
    'Description',
]

ROW_KEYS_PL = [
    'Tytuł',
    'Nazwisko',
    'Imię',
    'Wydawnictwo',
    'Miasto',
    'Rok',
    'Kategoria',
    'ISBN',
    'Wypożyczenia',
    'Dostępność',
    'Litera',
    'Biała kartka',
    'Kolorowa kartkaRodzaj',
    'Ilość',
    'Oznakowanie',
    'Adnotacja',
    'Opis',
]

DEBUG = False
NUMBER_OF_ROWS_INI = 1
NUMBER_OF_ROWS_END = None

def main():
    """
    """
    global NUMBER_OF_ROWS_INI, NUMBER_OF_ROWS_END
    # initiate XLSX
    wb = load_workbook('Katalog 20191020.xlsx')
    sheet = wb.active
    yaml_data = list()

    # Extract the header
    NUMBER_OF_ROWS_END = sheet.max_row if NUMBER_OF_ROWS_END is None else NUMBER_OF_ROWS_END
    for row_index in range(NUMBER_OF_ROWS_INI, NUMBER_OF_ROWS_END + 1):
        if DEBUG is True:
            print("\nRow number %d:" % row_index)
        row_excel = sheet[row_index]
        raw_row = [cell.value for cell in row_excel]

        # Build row dictionary
        row_dict = OrderedDict(zip(ROW_KEYS, raw_row))
        if isinstance(row_dict['Number'], float):
            row_dict['Number'] = int(row_dict['Number'])
        if isinstance(row_dict['Year'], float):
            row_dict['Year'] = int(row_dict['Year'])
        if isinstance(row_dict['Amount'], float):
            row_dict['Amount'] = int(row_dict['Amount'])

        if DEBUG is True:
            for key, value in row_dict.items():
                print("%-15s: %s" % (key, value))
            pdb.set_trace()

        # Attach to the yaml_data
        yaml_data.append(dict(row_dict))

    # dump the yaml_data into an actual data file
    file_name = 'Katalog_2019_10_20_%s_%s.yaml' % (NUMBER_OF_ROWS_INI, NUMBER_OF_ROWS_END)
    with open(file_name, 'w') as file_descriptor:
        yaml.dump(yaml_data, file_descriptor, indent=4)

    # header_xl = sheet[1]
    # for cell in header_xl:
        # print(cell.value)


if __name__ == "__main__":
    """
    """
    sys.exit(main())
